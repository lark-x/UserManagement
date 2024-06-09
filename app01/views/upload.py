import os
from datetime import datetime
from UserManagement import settings
from django.shortcuts import render, redirect, HttpResponse
from django.conf import settings
from app01 import models
from openpyxl import Workbook, load_workbook

from app01.utils.form import UploadForm, UploadModelForm, BillModelForm, ExcelFileModelForm

# 设置默认编码为UTF-8



def upload_list(request):
    if request.method == 'GET':
        return render(request, 'upload_list.html')
    if request.method == 'POST':
        file = request.FILES.get('avatar')
        f = open('avatar.jpg', 'wb')
        for chunk in file.chunks():
            f.write(chunk)
        f.close()
        return render(request, 'upload_list.html')


def upload_form(request):
    title = 'form上传'
    if request.method == 'GET':
        form = UploadForm()
        context = {
            'title': title, 'form': form
        }
        return render(request, 'upload_form.html', context)
    if request.method == 'POST':
        form = UploadForm(data=request.POST, files=request.FILES)
        # print(form.cleaned_data.values())
        if form.is_valid():
            # 1.读取图片内容,写入到文件夹中并获取文件的路径
            image_object = form.cleaned_data.get('img')

            media_path = os.path.join("media", image_object.name)
            # print(media_path)
            f = open(media_path, mode='wb')
            for chunk in image_object.chunks():
                f.write(chunk)
            f.close()
            # 2.将图片文件路径写入到数据库
            models.Boss.objects.create(
                name=form.cleaned_data.get('name'),
                age=form.cleaned_data.get('age'),
                img=media_path
            )
            return HttpResponse("...")
        return render(request, 'upload_form.html', {'form': form, 'title': title})


def upload_model_form(request):
    if request.method == 'GET':
        form = UploadModelForm()
        context = {
            'form': form,
            'title': 'ModelForm上传文件'
        }
        return render(request, 'upload_form.html', context)
    if request.method == 'POST':
        form = UploadModelForm(data=request.POST, files=request.FILES)
        context = {
            'form': form,
            'title': 'ModelForm上传文件'
        }
        if form.is_valid():
            # 自动保存数据
            form.save()
            redirect('/upload/modelform/')
        return render(request, 'upload_form.html', context)


def upload_excel(request):
    if request.method == 'GET':
        form = ExcelFileModelForm()
        context = {
            'form': form,
            'title': '账单列表'
        }
        return render(request, 'upload_excel.html', context)
    if request.method == 'POST':
        # 保存数据
        form = ExcelFileModelForm(request.POST, request.FILES)
        context = {
            'form': form,
            'title': '账单列表',
            'error': form.errors,
        }
        if form.is_valid():
            # 1.读取文件内容,写入到文件夹中并获取文件的路径
            # 2.将excel文件路径写入到数据库
            # media_path = os.path.join("media", image_object.name)
            form.instance.uploader_id = request.session['info']['id']
            form.save()
            # 获取excel文件路径
            file_path = os.path.join(settings.MEDIA_ROOT, str(form.instance.filePath))
            # bug修改
            # windows路径用\分隔,但是django中的modelform对fileField路径的存储使用的是/
            # 把路径字符串中的'/'全部替换成'\'
            file_path = (str(file_path)).replace('/', '\\')
            # 获取excel文件路径并且读取文件
            wb = load_workbook(file_path)
            ws = wb.active
            column_letter1, column_letter2, column_letter3, column_letter4 = 'A', 'A', 'A', 'A'
            row_start = 1
            for row in ws.rows:
                for cell in row:
                    if cell.value == '交易时间':
                        row_start = cell.row + 1
                        column_letter1 = cell.column
                    elif cell.value == '交易对方':
                        column_letter2 = cell.column
                    elif cell.value == '收/支':
                        column_letter3 = cell.column
                    elif cell.value == '金额(元)':
                        column_letter4 = cell.column
            for i in range(row_start, ws.max_row + 1):
                # bill = BillModelForm()
                datatime_str = ws.cell(row=i, column=column_letter1).value
                bill_datetime = datetime.strftime(datatime_str, '%Y-%m-%d %H:%M:%S')
                bill_description = ws.cell(row=i, column=column_letter2).value
                bill_type = ws.cell(row=i, column=column_letter3).value
                # amount = re.findall(r'\d+\.\d+', ws.cell(row=i, column=column_letter4).value)
                bill_amount = ws.cell(row=i, column=column_letter4).value
                bill_recorder_id = request.session['info']['id']
                # raw_query = 'insert into app01_bill(data,type,amount,description,recorder_id) values (
                # bill_datetime, bill_type, bill_amount, bill_description, bill_recorder_id)'
                models.Bill.objects.create(data=bill_datetime, description=bill_description, type=bill_type,
                                           amount=bill_amount, recorder_id=bill_recorder_id)
                # if bill.is_valid():
                #     print(bill.instance)
                #     print(1)
                #     bill.save()
            wb.close()
            # select_columns = [ws[column_letter1].value for i in range(row_start + 1, ws.max_row + 1)]
            # select_columns += [ws[column_letter2].value for i in range(row_start + 1, ws.max_row + 1)]
            # select_columns += [ws[column_letter3].value for i in range(row_start + 1, ws.max_row + 1)]
            # select_columns += [ws[column_letter4].value for i in range(row_start + 1, ws.max_row + 1)]
            return redirect('/upload/excel/')

            # 读取数据并进行处理
        return render(request, 'upload_excel.html', context)
